import os
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter

# ── Supabase connection ─────────────────────────────
SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL') or \
               os.environ.get('SUPABASE_URL')
SERVICE_KEY  = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')

if not SUPABASE_URL or not SERVICE_KEY:
    raise ValueError(
        'Missing env vars. Set SUPABASE_URL and '
        'SUPABASE_SERVICE_ROLE_KEY before running.'
    )

supabase = create_client(SUPABASE_URL, SERVICE_KEY)

# ── Fetch ALL scholarships ───────────────────────────
print('Fetching scholarships...')
response = supabase.table('scholarships') \
    .select('*') \
    .order('created_at', desc=False) \
    .execute()

scholarships = response.data
print(f'Found {len(scholarships)} scholarships')

if not scholarships:
    print('No scholarships found. Exiting.')
    exit(0)

# ── Column definitions ───────────────────────────────
COLUMNS = [
    ('id',                          'UUID (ห้ามแก้ไข)'),
    ('name_th',                     'ชื่อทุน (ภาษาไทย)*'),
    ('name_en',                     'ชื่อทุน (ภาษาอังกฤษ)'),
    ('funder_name_th',              'ผู้ให้ทุน (ภาษาไทย)*'),
    ('funder_name_en',              'ผู้ให้ทุน (ภาษาอังกฤษ)'),
    ('funder_type',                 'ประเภทผู้ให้ทุน'),
    ('amount_thb',                  'จำนวนเงิน (บาท)'),
    ('amount_type',                 'รูปแบบทุน'),
    ('is_loan',                     'เงินกู้ (TRUE/FALSE)*'),
    ('min_gpa',                     'เกรดขั้นต่ำ'),
    ('max_income_thb',              'รายได้สูงสุด (บาท/เดือน)'),
    ('welfare_card_priority',       'บัตรสวัสดิการ (TRUE/FALSE)'),
    ('grade_levels',                'ระดับการศึกษา'),
    ('field_of_study',              'สาขาวิชา'),
    ('province_restriction',        'จังหวัด'),
    ('enrolled_university_required','ต้องเป็นนักศึกษาของ'),
    ('english_level',               'ระดับภาษาอังกฤษ'),
    ('english_score_required',      'คะแนนภาษาที่ต้องการ'),
    ('bond_obligation',             'มีข้อผูกพัน (TRUE/FALSE)'),
    ('renewable',                   'ต่ออายุได้ (TRUE/FALSE)'),
    ('documents_required',          'เอกสารที่ต้องใช้'),
    ('description_th',              'คำอธิบาย (ภาษาไทย)'),
    ('deadline_date',               'วันปิดรับสมัคร*'),
    ('application_url',             'ลิงก์สมัคร*'),
    ('source_url',                  'ลิงก์แหล่งข้อมูล*'),
    ('historical_bias_score',       'Bias Score (วิจัย)'),
    ('is_active',                   'แสดงบนเว็บ (TRUE/FALSE)'),
    ('last_verified_at',            'ตรวจสอบล่าสุด'),
    ('created_at',                  'วันที่เพิ่ม'),
]

# ── Style helpers ───────────────────────────────────
NAVY       = '0A2342'
NAVY_MID   = '1B3A6B'
LIGHT_BLUE = 'EBF2FF'
WHITE      = 'FFFFFF'

def fill(hex_col):
    return PatternFill('solid', fgColor=hex_col)

def border():
    s = Side(style='thin', color='DDE3EE')
    return Border(left=s, right=s, top=s, bottom=s)

# ── Build workbook ───────────────────────────────────
wb = Workbook()

# ── Sheet 1: All Scholarships ────────────────────────
ws = wb.active
ws.title = 'Scholarships (ALL)'

for col_idx, (field, label) in enumerate(COLUMNS, start=1):
    c1 = ws.cell(row=1, column=col_idx, value=field)
    c1.fill = fill(NAVY)
    c1.font = Font(name='Arial', bold=True, color=WHITE, size=10)
    c1.alignment = Alignment(vertical='center')
    c1.border = border()

    c2 = ws.cell(row=2, column=col_idx, value=label)
    c2.fill = fill(LIGHT_BLUE)
    c2.font = Font(name='Arial', italic=True, color=NAVY_MID, size=9)
    c2.alignment = Alignment(vertical='center', wrap_text=True)
    c2.border = border()

for row_idx, s in enumerate(scholarships, start=3):
    for col_idx, (field, _) in enumerate(COLUMNS, start=1):
        raw = s.get(field)

        if isinstance(raw, list):
            val = ', '.join(str(x) for x in raw)
        elif raw is None:
            val = ''
        else:
            val = raw

        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=10, color='0A2342')
        cell.alignment = Alignment(vertical='center')
        cell.border = border()

        if field == 'is_active' and raw is False:
            cell.fill = fill('FEE2E2')
        elif row_idx % 2 == 0:
            cell.fill = fill('F9FAFB')

widths = {
    'id': 38, 'name_th': 40, 'name_en': 32,
    'funder_name_th': 32, 'funder_name_en': 28,
    'funder_type': 14, 'amount_thb': 14,
    'amount_type': 12, 'is_loan': 10,
    'min_gpa': 10, 'max_income_thb': 18,
    'welfare_card_priority': 16, 'grade_levels': 18,
    'field_of_study': 28, 'province_restriction': 24,
    'enrolled_university_required': 30,
    'english_level': 14, 'english_score_required': 20,
    'bond_obligation': 14, 'renewable': 14,
    'documents_required': 35, 'description_th': 45,
    'deadline_date': 14, 'application_url': 38,
    'source_url': 38, 'historical_bias_score': 16,
    'is_active': 12, 'last_verified_at': 20,
    'created_at': 20,
}
for col_idx, (field, _) in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(field, 18)

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 44
ws.freeze_panes = 'A3'

# ── Sheet 2: Summary ─────────────────────────────────
ws2 = wb.create_sheet('Summary')

active_count   = sum(1 for s in scholarships if s.get('is_active'))
inactive_count = sum(1 for s in scholarships if not s.get('is_active'))

funder_counts = Counter(
    s.get('funder_type', 'unknown') for s in scholarships
)

summary_rows = [
    ('NEWEST MASTERSHEET — TunDee ทุนดี', True),
    (f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M")}', False),
    ('', False),
    ('COUNTS', True),
    (f'Total scholarships: {len(scholarships)}', False),
    (f'Active (shown on site): {active_count}', False),
    (f'Inactive (hidden): {inactive_count}', False),
    ('', False),
    ('BY FUNDER TYPE', True),
]

for ftype, count in sorted(funder_counts.items()):
    summary_rows.append((f'  {ftype}: {count}', False))

summary_rows += [
    ('', False),
    ('NOTE', True),
    ('All scholarship data has been set to is_active=FALSE', False),
    ('on tundee.org. Scholarships are NOT deleted from', False),
    ('the database — they are only hidden from public view.', False),
    ('To restore: run UPDATE scholarships SET is_active=TRUE', False),
    ('in Supabase SQL Editor.', False),
]

for r_idx, (text, is_header) in enumerate(summary_rows, start=1):
    cell = ws2.cell(row=r_idx, column=2, value=text)
    if is_header and text:
        cell.fill = fill(NAVY)
        cell.font = Font(name='Arial', bold=True, color=WHITE, size=11)
    else:
        cell.font = Font(name='Arial', size=10, color='0A2342')
    ws2.row_dimensions[r_idx].height = 18

ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 55

# ── Save ─────────────────────────────────────────────
output_path = 'NEWEST_MASTERSHEET.xlsx'
wb.save(output_path)
print(f'\nSaved: {output_path}')
print(f'Scholarships exported: {len(scholarships)}')
print(f'Active: {active_count}  |  Inactive: {inactive_count}')
print('Done.')
